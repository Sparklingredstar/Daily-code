# -*- coding: utf-8 -*-
"""
Created on Tue Oct  8 10:40:40 2019

@author: adrian
"""
import openpyxl

wb = openpyxl.load_workbook('子牙河.xlsx',data_only=True)
sheet = wb['Sheet1']
wb1 = openpyxl.load_workbook('子牙河统计.xlsx')
sheet1 = wb1['Sheet1']

count1 = 3


for row1 in range(2, 47):
    h_value = sheet['H' + str(row1)].value
    i_value = sheet['I' + str(row1)].value
    high_value = sheet['E' + str(row1)].value
    w_value = sheet['D' + str(row1)].value
    if (float(h_value) < 0 and float(i_value) > 0 ):
        sheet1.cell(row = count1,column = 4, value = "桥梁净宽") 
        sheet1.cell(row = count1,column = 1, value = sheet.cell(row = row1,column = 1).value)
        sheet1.cell(row = count1,column = 2, value = sheet.cell(row = row1,column = 2).value)
        sheet1.cell(row = count1,column = 3, value = sheet.cell(row = row1,column = 3).value)
        sheet1.cell(row = count1,column = 8, value = "——")
        sheet1.cell(row = count1,column = 6, value = h_value)
        count1 += 1
        print(count1)
    if (float(h_value) > 0 and float(i_value) < 0 ):
        sheet1.cell(row = count1,column = 4, value = "桥梁净高")
        sheet1.cell(row = count1,column = 1, value = sheet.cell(row = row1,column = 1).value)
        sheet1.cell(row = count1,column = 2, value = sheet.cell(row = row1,column = 2).value)
        sheet1.cell(row = count1,column = 3, value = sheet.cell(row = row1,column = 3).value)
        sheet1.cell(row = count1,column = 6, value = "——")
        sheet1.cell(row = count1,column = 8, value = i_value)
        count1 += 1
        
    if (float(h_value) < 0 and float(i_value) < 0 ):
        sheet1.cell(row = count1,column = 4, value = "桥梁净宽、净高")
        sheet1.cell(row = count1,column = 1, value = sheet.cell(row = row1,column = 1).value)
        sheet1.cell(row = count1,column = 2, value = sheet.cell(row = row1,column = 2).value)
        sheet1.cell(row = count1,column = 3, value = sheet.cell(row = row1,column = 3).value)
        sheet1.cell(row = count1,column = 6, value = h_value)
        sheet1.cell(row = count1,column = 8, value = i_value)
        count1 += 1  
    print(count1)
    wb1.save('子牙河统计.xlsx')


    
        


 